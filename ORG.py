import requests
from bs4 import BeautifulSoup
from requests.api import head
import re
import time
import openpyxl
from datetime import date
today = date.today().strftime("%d-%m-%Y")

def find_row(sheet, product_title, product_id):
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == product_title and sheet.cell(row, 2).value == product_id:
            return row
    return None

wb = openpyxl.load_workbook(filename='ORG Прайс.xlsx')
sheet = wb.active
header = ['Title', 'ID']
if f"Price({today})" not in [cell.value for cell in sheet[1]]:
    sheet.cell(row=1, column=sheet.max_column + 1).value = f"Price({today})"
    header.append(f"Price({today})")

with open('ORG ID_product.txt', 'r') as f:
    data = f.read().splitlines()
    for i in range(0, len(data), 10):

        list_id = ','.join(data[i:i+10])
        s = requests.Session()
        url = f'http://ORG.com/compare/{list_id}/'
        print(url)
        response = s.get(url=url)
        time.sleep(5)
        soup = BeautifulSoup(response.content, 'html.parser')
        x = soup.find_all('table', class_='compare')
        #print(x)
        for i in x:
            title_list = i.find_all('a', class_='image-link')
            price_list = i.find_all('span', class_ = 'price nowrap')
            for title, price in zip(title_list, price_list):
                img = title.find('img', itemprop='image')
                product_title = img.get('alt')
                match = re.search(r'/(\d+)/images', img.get('src'))
                product_id = int(match.group(1)) if match else None
                product_price = float(price.get_text().replace(',', '.').replace('Р', '').replace(' ', ''))
                print(product_title, product_id, product_price)
                row = find_row(sheet, product_title, product_id)
                if row:
                    col = header.index(f"Price({today})") + 1
                    sheet.cell(row, col).value = product_price
                else:
                    row = sheet.max_row + 1
                    sheet.append([product_title, product_id, product_price])

                time.sleep(1)

wb.save("ORG Прайс.xlsx")

